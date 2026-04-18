import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, queue, time, os, configparser
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import traceback

default_font = Font(name='Microsoft YaHei', size=11)
header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
total_font = Font(name='Microsoft YaHei', size=11, bold=True)

CONFIG_FILE = "config.ini"

# ================= CONFIG =================
def load_config():
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE, encoding='utf-8')
    else:
        config['PATH'] = {'input': '', 'output': ''}
        config['TIME'] = {'start': '12:00', 'end': '12:00'}
    return config

def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        config.write(f)
def apply_heatmap_safe(ws, enable):
    if not enable:
        return

    # ===== HEADER MAP =====
    headers = [str(cell.value).strip() for cell in ws[1]]
    col_map = {name: idx+1 for idx, name in enumerate(headers)}

    shift_a_col = col_map.get("Total Shift A")
    shift_b_col = col_map.get("Total Shift B")
    grand_col   = col_map.get("Grand Total")

    # debug (เปิดไว้ตอนทดสอบ)
    print("MAP:", col_map)

    # ถ้า column ไม่ครบ = skip
    if not (shift_a_col and shift_b_col and grand_col):
        print("❌ Skip heatmap: column not found")
        return

    max_row = ws.max_row - 1
    start_col = 2  # B

    # ===== COLOR RULE =====
    # 🟢 Shift A → เขียว (3 ระดับ)
    rule_a = ColorScaleRule(
        start_type='min', start_color='F8696B',      # แดง (น้อย)
        mid_type='percentile', mid_value=50, mid_color='FFEB84',  # เหลือง
        end_type='max', end_color='63BE7B'           # เขียว (มาก)
    )

    # 🔵 Shift B → ฟ้า (3 ระดับ)
    rule_b = ColorScaleRule(
        start_type='min', start_color='F8696B',      
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'           
    )

    # ===== SAFE RANGE APPLY =====

    def safe_range(col_start, col_end):
        if col_start and col_end and col_end >= col_start:
            return f"{get_column_letter(col_start)}2:{get_column_letter(col_end)}{max_row}"
        return None

    # A zone
    range_a = safe_range(start_col, shift_a_col - 1)

    # B zone
    range_b = safe_range(shift_a_col + 1, shift_b_col - 1)

    # C zone
    range_c = safe_range(shift_b_col + 1, grand_col - 1)

    # ===== APPLY =====
    if range_a:
        ws.conditional_formatting.add(range_a, rule_a)

    if range_b:
        ws.conditional_formatting.add(range_b, rule_b)

    if range_c:
        ws.conditional_formatting.add(range_c, rule_b)

    print("✅ Heatmap applied safely")
# ================= APP =================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Scan Report Tool")
        self.root.geometry("720x720")
        self.root.resizable(False, False)

        self.queue = queue.Queue()
        self.running = False
        self.start_time = 0

        self.config = load_config()

        self.input_path = tk.StringVar(value=self.config['PATH'].get('input', ''))
        self.start_hour = tk.StringVar(value=self.config['TIME'].get('start', '12'))
        self.end_hour = tk.StringVar(value=self.config['TIME'].get('end', '12'))

        self.enable_category = tk.BooleanVar(
            value=self.config.get('OPTION', 'category', fallback='True') == 'True'
        )

        self.enable_heatmap = tk.BooleanVar(
            value=self.config.get('OPTION', 'heatmap', fallback='True') == 'True'
        )
        
        #heatmap_default = self.config.get('OPTION', 'heatmap', fallback='True') == 'True'
        #self.enable_heatmap = tk.BooleanVar(value=heatmap_default)

        self.build_ui()
        self.root.after(100, self.process_queue)

    # ================= UI =================
    def build_ui(self):
        # Header
        header = tk.Frame(self.root, bg="#2f4f6f", height=50)
        header.pack(fill="x")
        tk.Label(header, text="Scan Report Processor",
                 bg="#2f4f6f", fg="white",
                 font=("Segoe UI", 16, "bold")).pack(pady=10)

        main = tk.Frame(self.root, padx=10, pady=10)
        main.pack(fill="both", expand=True)

        # ===== Input =====
        box1 = ttk.LabelFrame(main, text="Input File", padding=10)
        box1.pack(fill="x", pady=5)

        box1.columnconfigure(0, weight=1)  # ✅ ให้ช่อง input ขยาย

        self.entry_input = ttk.Entry(box1, textvariable=self.input_path)
        self.entry_input.grid(row=0, column=0, sticky="ew", padx=(0,5))

        self.btn_browse = ttk.Button(box1, text="Browse", command=self.browse_input)
        self.btn_browse.grid(row=0, column=1)

        # ===== Progress =====
        box2 = ttk.LabelFrame(main, text="Progress", padding=10)
        box2.pack(fill="x", pady=5)

        self.progress = ttk.Progressbar(box2, length=650, mode='determinate')
        self.progress.pack(pady=5)

        self.status_label = tk.Label(box2, text="0%", anchor="w")
        self.status_label.pack(fill="x")

        self.eta_label = tk.Label(box2, text="ETA: -", anchor="e")
        self.eta_label.pack(fill="x")

        # ===== Log =====
        box3 = ttk.LabelFrame(main, text="Log", padding=5)
        box3.pack(fill="both", expand=True, pady=5)

        self.log = tk.Text(box3, height=12)
        self.log.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(box3, command=self.log.yview)
        scrollbar.pack(side="right", fill="y")
        self.log.config(yscrollcommand=scrollbar.set)

        # ===== Bottom =====
        bottom = tk.Frame(main)
        bottom.pack(fill="x", pady=10)

        # Export Setting
        Setting_frame = ttk.LabelFrame(bottom, text="Export Setting", padding=10)
        Setting_frame.pack(side="left", padx=10)

        ttk.Label(Setting_frame, text="Time Range").grid(row=0, column=0, padx=5)

        self.cb_start = ttk.Combobox(
            Setting_frame,
            textvariable=self.start_hour,
            values=[f"{i:02d}:00" for i in range(24)],
            width=6,
            state="readonly"
        )
        self.cb_start.grid(row=0, column=1)

        ttk.Label(Setting_frame, text="→").grid(row=0, column=2, padx=5)

        self.cb_end = ttk.Combobox(
            Setting_frame,
            textvariable=self.end_hour,
            values=[f"{i:02d}:00" for i in range(24)],
            width=6,
            state="readonly"
        )
        self.cb_end.grid(row=0, column=3)

        # Category
        self.chk_category = ttk.Checkbutton(
            Setting_frame,
            text="Category Color",
            variable=self.enable_category
        )
        self.chk_category.grid(row=0, column=4, padx=5)

        # Heatmap
        self.chk_heatmap = ttk.Checkbutton(
            Setting_frame,
            text="Heatmap",
            variable=self.enable_heatmap
        )
        self.chk_heatmap.grid(row=0, column=5, padx=5)

        tk.Label(header,
            text="Developed by Siwamon Singtan : IT KKN | V1.1",
            bg="#2f4f6f", fg="white",
            font=("Segoe UI", 8)
        ).pack()

        # Button (ซ้อน)
        self.btn_frame = tk.Frame(bottom)
        self.btn_frame.pack(side="right")

        self.btn_process = tk.Button(self.btn_frame,
                                     text="▶ PROCESS",
                                     bg="#35c16c", fg="white",
                                     font=("Segoe UI", 12, "bold"),
                                     width=20, height=2,
                                     command=self.start_process)

        self.btn_cancel = tk.Button(self.btn_frame,
                                    text="✖ CANCEL",
                                    bg="#d9534f", fg="white",
                                    font=("Segoe UI", 12, "bold"),
                                    width=20, height=2,
                                    command=self.cancel_process)

        self.btn_process.grid(row=0, column=0)
        self.btn_cancel.grid(row=0, column=0)
        self.btn_cancel.grid_remove()

    def set_ui_state(self, state):
        self.entry_input.config(state=state)
        self.btn_browse.config(state=state)
        self.chk_category.config(state=state)
        self.chk_heatmap.config(state=state)
        self.cb_start.config(state="readonly" if state == "normal" else "disabled")
        self.cb_end.config(state="readonly" if state == "normal" else "disabled")

    # ================= UI FUNC =================
    def browse_input(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.input_path.set(path)

    def log_message(self, msg):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)

    def process_queue(self):
        try:
            while True:
                msg, val = self.queue.get_nowait()

                if msg == "log":
                    self.log_message(val)

                elif msg == "progress":
                    self.progress['value'] = val
                    self.status_label.config(text=f"{val}%")

                    if self.running and val > 0:
                        elapsed = time.time() - self.start_time
                        eta = (elapsed / val) * (100 - val)
                        self.eta_label.config(text=f"ETA: {eta:.1f}s")

                elif msg == "done":
                    self.finish()

        except:
            pass

        self.root.after(100, self.process_queue)

    # ================= PROCESS =================
    def start_process(self):
        if not self.input_path.get():
            messagebox.showerror("Error", "Select input file")
            return

        self.running = True
        self.start_time = time.time()
        self.progress['value'] = 0

        self.btn_process.grid_remove()
        self.btn_cancel.grid()

        self.set_ui_state("disabled")

        threading.Thread(target=self.run_process, daemon=True).start()

    def cancel_process(self):
        self.running = False
        self.set_ui_state("normal")

        self.btn_cancel.grid_remove()
        self.btn_process.grid()

        self.progress['value'] = 0

        self.queue.put(("log", "❌ Cancelled"))

    def finish(self):
        self.set_ui_state("normal")
        self.running = False
        self.btn_cancel.grid_remove()
        self.btn_process.grid()
        messagebox.showinfo("Done", "Process Complete")

    # ================= CORE LOGIC =================
    def run_process(self):
        try:
            self.queue.put(("log", "📥 Load file..."))
            df = pd.read_excel(self.input_path.get())
            self.queue.put(("progress", 10))

            self.queue.put(("log", "🧹 Prepare data..."))
            df['scan_time'] = pd.to_datetime(df.iloc[:, 3], errors='coerce')
            df['Station'] = df.iloc[:, 5]

            try:
                file_date = df['scan_time'].dropna().iloc[0].strftime('%Y-%m-%d')
            except:
                file_date = time.strftime('%Y-%m-%d')

            # เก็บไว้ใช้ตอน save
            self.file_date = file_date

            self.queue.put(("progress", 25))

            self.queue.put(("log", "⏱ Generate time slot..."))
            def get_time_slot(dt):
                hour = dt.hour
                if hour < 12:
                    hour += 24
                return f"{hour%24:02d}.00-{(hour+1)%24:02d}.00"

            df['time_slot'] = df['scan_time'].apply(get_time_slot)
            self.queue.put(("progress", 40))

            self.queue.put(("log", "📊 Pivot..."))
            result = df.groupby(['Station', 'time_slot']).size().unstack(fill_value=0)
            self.queue.put(("progress", 55))

            # ===== Time Range =====
            start = int(self.start_hour.get().split(":")[0])
            end = int(self.end_hour.get().split(":")[0])

            hours = [(start + i) % 24 for i in range(24)]
            """time_order = [f"{h:02d}.00-{(h+1)%24:02d}.00" for h in hours]

            if start == end:
                selected = time_order
            else:
                idx = hours.index(end)
                selected = time_order[:idx]"""
            time_order = [f"{h:02d}.00-{(h+1)%24:02d}.00" for h in hours]
            if start == end:
                length = 24
            else:
                length = (end - start) % 24
                if length == 0:
                    length = 24

            selected = time_order[:length]

            result = result.reindex(columns=selected, fill_value=0).reset_index()

            # ===== Sort =====
            def sort_key(name):
                if isinstance(name, str):
                    if name.endswith('_GW'):
                        return (0, name)
                    elif name.startswith('D_'):
                        return (1, name)
                    elif name and name[0].isdigit():
                        return (2, name)
                return (3, name)

            result = result.sort_values(
                by='Station',
                key=lambda col: col.map(sort_key)
            ).reset_index(drop=True)

           # ===== SHIFT CALCULATION =====
            time_cols = result.columns[1:]

            shift_a_cols = []
            shift_b_cols = []

            for col in time_cols:
                start_hour = int(col[:2])

                if start_hour >= 12 or start_hour == 0:
                    shift_a_cols.append(col)
                else:
                    shift_b_cols.append(col)

            result['Total Shift A'] = result[shift_a_cols].sum(axis=1)
            result['Total Shift B'] = result[shift_b_cols].sum(axis=1)
            result['Grand Total'] = result['Total Shift A'] + result['Total Shift B']

            # ===== TOTAL ROW =====
            time_only = result.columns[1:-3]

            total_row = result[time_only].sum()
            total_row['Total Shift A'] = result['Total Shift A'].sum()
            total_row['Total Shift B'] = result['Total Shift B'].sum()
            total_row['Grand Total'] = result['Grand Total'].sum()
            total_row['Station'] = 'TOTAL ALL STATIONS'

            result = pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)

            # ===== REORDER COLUMN =====

            cols = list(result.columns)

            # หา index ของ 00.00-01.00
            target_col = None
            for c in cols:
                if c.startswith("00.00-01.00"):
                    target_col = c
                    break

            if target_col:
                cols.remove('Total Shift A')
                insert_idx = cols.index(target_col) + 1
                cols.insert(insert_idx, 'Total Shift A')

                # จัดตำแหน่งใหม่
                result = result[cols]

            self.queue.put(("progress", 70))

            # ===== SAVE AFTER PROCESS =====

            # เอาเฉพาะ column เวลาเท่านั้น
            time_cols = result.columns[1:-3]  # ตัด shift + grand total ออก
            total_scan = int(result['Grand Total'].iloc[-1])

            self.queue.put(("log", f"📦 Total Scan: {total_scan:,}"))

            self.queue.put(("log", "💾 Waiting for save..."))
            self.root.after(0, self.ask_save, result)

        except Exception as e:
            self.queue.put(("log", traceback.format_exc()))

    def ask_save(self, result):
        initial_dir = self.config['PATH'].get('output', '')
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialdir=initial_dir,
            initialfile=f"สแกนบรรจุขึ้นรถ_{self.file_date}.xlsx"
        )
        if 'OPTION' not in self.config:
            self.config['OPTION'] = {}

        self.config['OPTION']['heatmap'] = str(self.enable_heatmap.get())
        self.config['OPTION']['category'] = str(self.enable_category.get())

        if not save_path:
            self.queue.put(("log", "❌ Cancel save"))
            self.finish()
            return

        self.queue.put(("log", "💾 Saving..."))

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            result.to_excel(writer, index=False, sheet_name="Report")
            ws = writer.sheets["Report"]
            # ===== STYLE =====
            #header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")  # น้ำเงิน
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ===== HEADER =====
            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

                if "Total Shift A" in cell.value:
                    cell.fill = PatternFill("solid", fgColor="F4B084")

                if "Total Shift B" in cell.value:
                    cell.fill = PatternFill("solid", fgColor="F4B084")

                if "Grand Total" in cell.value:
                    cell.fill = PatternFill("solid", fgColor="A9D08E")
            # ===== DATA =====
            header_names = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2):

                is_total = str(row[0].value) == 'TOTAL ALL STATIONS'

                for col_num, cell in enumerate(row, 1):

                    col_name = header_names[col_num-1]

                    cell.font = default_font
                    cell.border = thin_border
                
                    # ===== ALIGN =====
                    if col_num == 1:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

                    # ===== NUMBER FORMAT =====
                    if isinstance(cell.value, (int, float)) and col_num != 1:
                        cell.number_format = '#,##0'

                    # ===== CATEGORY COLOR =====
                    if col_num == 1 and self.enable_category.get() and not is_total:
                        name = str(cell.value)

                        if name.endswith('_GW'):
                            cell.fill = PatternFill("solid", fgColor="95B3D7")
                        elif name.startswith('D_'):
                            cell.fill = PatternFill("solid", fgColor="B8CCE4")
                        elif name and name[0].isdigit():
                            cell.fill = PatternFill("solid", fgColor="DCE6F1")

                    # ===== LOCK SHIFT / TOTAL (ขาว) =====
                    if col_name in ["Total Shift A", "Total Shift B", "Grand Total"]:
                        cell.fill = PatternFill("solid", fgColor="FFFFFF")

                    # ===== TOTAL ROW =====
                    if is_total:
                        cell.font = total_font
                        cell.fill = PatternFill("solid", fgColor="D9D9D9")

            # ===== FREEZE + FILTER =====
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # ===== COLOR SCALE (Heatmap) =====
            apply_heatmap_safe(ws, self.enable_heatmap.get())

            # ===== AUTO WIDTH (ทำทุกครั้ง) =====
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+2,40)

        # save config
        self.config['PATH']['input'] = self.input_path.get()
        self.config['PATH']['output'] = os.path.dirname(save_path)
        self.config['TIME']['start'] = self.start_hour.get()
        self.config['TIME']['end'] = self.end_hour.get()
        save_config(self.config)

        self.queue.put(("progress", 100))
        self.queue.put(("log", f"✅ Saved → {save_path}"))
        self.queue.put(("done", None))


# ================= RUN =================
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()